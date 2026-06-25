"""
Write agent results back into the single-vendor scorecard template.

The main Scorecard sheet keeps the original template layout. A second sheet,
"Consistency Check", captures the smarter judge review so the Excel file shows
where the judge agreed, disagreed, or recommended score adjustments.
"""

from __future__ import annotations

import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import config

YELLOW = PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
HEADER_FILL = PatternFill("solid", start_color="FFDBEAFE", end_color="FFDBEAFE")


def _safe_sheet_name(name: str, existing: set[str]) -> str:
    candidate = name[:31]
    if candidate not in existing:
        return candidate
    i = 2
    while True:
        suffix = f" {i}"
        candidate = f"{name[:31 - len(suffix)]}{suffix}"
        if candidate not in existing:
            return candidate
        i += 1


def _write_consistency_sheet(wb, state: dict) -> None:
    sheet_name = "Consistency Check"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(_safe_sheet_name(sheet_name, set(wb.sheetnames)))

    consistency = state.get("consistency")

    ws["A1"] = "Judge Consistency Check"
    ws["A1"].font = Font(bold=True, size=14)

    if consistency is None:
        ws["A3"] = "No consistency check was available for this run."
        return

    summary_rows = [
        ("Judge model", getattr(consistency, "judge_model", "")),
        ("Agreement within 1 point", getattr(consistency, "agreement_within_1", 0.0)),
        ("Mean absolute difference", getattr(consistency, "mean_abs_diff", 0.0)),
        ("Material judge findings", getattr(consistency, "material_issues_count", 0)),
        ("Summary", getattr(consistency, "note", "")),
    ]

    row = 3
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = [
        "Criterion",
        "Issue Type",
        "Severity",
        "Primary Score",
        "Judge Score",
        "Delta",
        "Recommended Score",
        "Material?",
        "Explanation",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    findings = getattr(consistency, "judge_findings", []) or []
    if not findings:
        row += 1
        ws.cell(row=row, column=1, value="No material judge findings.")
    else:
        for finding in findings:
            row += 1
            ws.cell(row=row, column=1, value=finding.get("description") or finding.get("criterion_id"))
            ws.cell(row=row, column=2, value=finding.get("issue_type"))
            ws.cell(row=row, column=3, value=finding.get("severity"))
            ws.cell(row=row, column=4, value=finding.get("primary_score"))
            ws.cell(row=row, column=5, value=finding.get("judge_score"))
            ws.cell(row=row, column=6, value=finding.get("delta"))
            ws.cell(row=row, column=7, value=finding.get("recommended_score"))
            ws.cell(row=row, column=8, value="Yes" if finding.get("material") else "No")
            ws.cell(row=row, column=9, value=finding.get("explanation"))

    widths = {
        "A": 42,
        "B": 24,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 10,
        "G": 18,
        "H": 12,
        "I": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def populate_template(template_path: str, out_path: str, state: dict) -> str:
    wb = load_workbook(template_path)
    ws = wb["Scorecard"]

    # Vendor name lives just under the title band in the template.
    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            if cell.value == "Vendor Name":
                cell.value = f"Vendor Name: {state['vendor_name']}"
                break

    scores = {s.criterion_id: s for s in state["scores"]}
    desc_to_id = {c.description: c.id for c in config.CRITERIA}

    # Walk rows; criterion rows have a description in column B.
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        desc_cell = row[1]  # column B
        cid = desc_to_id.get(desc_cell.value)
        if not cid:
            continue

        s = scores.get(cid)
        if not s:
            continue

        score_cell = row[3]  # column D — Score (1-5)
        src_cell = row[5]  # column F — Source
        score_cell.value = s.score if s.score is not None else None  # blank, not 0

        note = []
        if s.capped:
            note.append("[tier-capped]")
        if s.is_gap:
            note.append("[no evidence]")

        src = "; ".join(s.sources[:2])
        src_cell.value = (" ".join(note) + " " + src).strip() or (
            "[no evidence]" if s.is_gap else ""
        )

    _write_consistency_sheet(wb, state)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
